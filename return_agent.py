# File path manipulation, directory creation, env access
import os

# System-level access, here to add stdout as logging handler
import sys

# Provides time.sleep() for delays -> bot-detection avoidance
import time

# Random numbers for variable delays using random.uniform(min, max)
import random

# For logging. Better than print here b'coz it supports:
#   – log levels (DEBUG, INFO, WARNING, ERROR)
#   – simultaneous o/p to file & console
#   – timestamps
#   – named loggers for filtering
import logging

# For generating ISO 8601 timestamps
from datetime import datetime

# Decorator that auto-generates __init__, __repr__, __eq__ methods for data-holding classes.
# Reduces boilerplate significantly. 'field' allows setting default values for mutable types.
from dataclasses import dataclass, field

# Defines a set of named constants. Used for return status values to prevent typos and enable IDE autocomplete
from enum import Enum

# Library for reading/writing excel (.xlsx) files.
# Chosen over pandas b'coz:
#   – openpyxl preserves existing formatting
#   – allows cell-level writes
#   – safer for concurrent access (read-modify-write pattern)
import openpyxl

import argparse


# –––––––––––––
# CONFIGURATION
# –––––––––––––
# All config is defined as module-level constants.
# This makes it easy to find and modify settings without searching through the code.

# Path to the excel file containing return tasks.
# os.path.abspath(__file__) gets the absolute path to the script itself
# os.path.dirname() strips the filename, leaving the directory path
# os.path.join() constructs the full path in an OS-independent way
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Faym Status Test Orders.xlsx')

# Directory where screenshots are saved when the agent encounters issues.
SCREENSHOTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'screenshots')

# Directory for log files. Each run creates timestamped log file.
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')

FLIPKART_LOGIN_URL = "https://www.flipkart.com/account/login"
FLIPKART_ORDERS_URL = "https://www.flipkart.com/account/orders"

# Test Flipkart Account
FLIPKART_PHONE = "9205359199"

# Bot-detection avoidance timing config
MIN_ACTION_DELAY = 1.0
MAX_ACTION_DELAY = 3.5

MIN_TYPING_DELAY = 0.5
MAX_TYPING_DELAY = 0.15

# How long to wait for a page to load before timing out.
PAGE_LOAD_TIMEOUT = 30

MAX_RETRIES = 3

# Column mappings for the Excel file (1-indexed to match openpyxl convention).
# This class documents which column contains which data, preventing magic numbers scattered throughout the code.
# If the Excel structure changes, only this class needs to be updated.
class ExcelColumns:
    ADDRESS = 1         # Column A: Shipping address
    CONTACT = 2         # Column B: Contact phone number
    PRODUCT_LINK = 3    # Column C: Product URL(s) on Flipkart
    AMOUNT = 4          # Column D: Total order amount in ₹
    NUM_PRODUCTS = 5    # Column E: Number of products in the order
    ORDER_DATE = 6      # Column F: Date the order was placed
    ORDER_ID = 7        # Column G: Flipkart order ID (e.g., OD337915012166989100)
    DELIVERY_DATE = 8   # Column H: Expected/actual delivery date
    RETURN_WINDOW = 9   # Column I: Return eligibility window (e.g., "10 Days")
    STATUS = 10         # Column J: Task status (Pending/Done/Needs Review)
    PLATFORM = 11       # Column K: E-commerce platform (Flipkart/Amazon)
    REFUND_ID = 12      # Column L: Refund/Return ID assigned by the platform
    RETURN_STATUS = 13  # Column M: Detailed return status
    REFUND_AMOUNT = 14  # Column N: Amount refunded by the platform
    TIMESTAMP = 15      # Column O: When the agent last processed this row
    LOG = 16            # Column P: Detailed log/error message


# –––––––––––
# DATA MODELS
# –––––––––––
# These classes define the data structures used throughout the agent.
# Using @dataclass and Enum provides type safety and self-documentation.

class ReturnStatus(Enum):
    """
    Enumeration of all possible return statuses.
    Using an Enum prevents typos (e.g. "Pendingg") that would cause bugs,
    and enables IDE autocomplete
    """
    PENDING = "Pending"                           # Not yet processed
    PLACED = "Return Placed"                      # Return successfully initiated
    FAILED = "Failed"                             # Return attempt failed
    OUT_OF_WINDOW = "Out of Return Window"        # Past the return deadline
    NOT_DELIVERED = "Not yet delivered"            # Order hasn't arrived yet
    ALREADY_CANCELLED = "Already Cancelled & Refunded"  # Already handled
    NEEDS_REVIEW = "Support Needed"               # Requires manual intervention
    DONE = "Done"

@dataclass
class ReturnTask:
    """
    Represents a single return task (one line item) read from the Excel file.
    
    Each row in the Excel sheet becomes one ReturnTask instance.
    The row_number field is crucial - it tells us which excel row to write
    results back to after processing.
    
    Fields marked 'datatype | None' can be None when the column is empty in excel.
    This happens for new/unprocessed tasks where result columns haven't been filled yet.
    """
    row_number: int                        # Excel row number (2-indexed, since row 1 is headers)
    address: str                           # Shipping address of the customer
    contact: str                           # Contact phone number
    product_link: str                      # Flipkart product URL(s)
    amount: float                          # Order amount in ₹
    num_products: int                      # Number of products in the order
    order_date: str                        # When the order was placed
    order_id: str                          # Flipkart order ID (unique identifier)
    delivery_date: str                     # When the order was delivered
    return_window: str                     # Return eligibility window (e.g., "10 Days")
    status: str                            # Current task status
    platform: str                          # E-commerce platform (e.g., "Flipkart")
    refund_id: str | None = None        # Refund ID (filled by agent or pre-existing)
    return_status: str | None = None    # Detailed return status
    refund_amount: float | None = None  # Refund amount in ₹
    timestamp: str | None = None        # Last processing timestamp
    log: str | None = None              # Log/error message from previous runs


@dataclass
class ReturnResult:
    """
    The outcome of processing a single return task.
    
    This is what the agent produces after attempting a return.
    It's separate from ReturnTask because a task (input) and its result (output)
    are conceptually different - the task describes WHAT to do, the result
    describes WHAT HAPPENDED.
    """
    success: bool                           # True if return was successfully placed
    return_id: str | None = None            # Return/Refund ID from the platform
    return_status: str = ""                 # Human-readable status description
    refund_amount: float | None = None      # Refund amount in ₹ (from platform)
    log_message: str = ""                   # Detailed log for debugging
    screenshot_path: str | None = None      # Path to screenshot (for failed cases)


# –––––––––––––
# LOGGING SETUP
# –––––––––––––

def setup_logging():
    """
    Configure the logging system with dula ouput (file + console).
    
    Each run creates a new timestamped log file in the logs/ directory.
    This ensures:
        1. Console output for real-time monitoring during execution
        2. Persistent log files for post-mortem debugging
        3. Consistent timestamp format across all log entries
    
    Returns:
        logging.Logger: Configured logger instance named 'ReturnAgent'
    """
    os.makedirs(LOG_DIR, exist_ok=True)

    log_file = os.path.join(LOG_DIR, f"return_agent_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

    # Configure root logger with:
    #   - INFO level: logs INFO, WARNING, ERROR, CRITICAL (not DEBUG)
    #   - Format: timestamp [LEVEL] logger_name: message
    #   - Two handlers: file (persistent) and stdout (real-time)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger('ReturnAgent')


# ------------------------
# HUMAN BEHAVIOR SIMULATOR
# ------------------------

class HumanSimulator:
    """
    Simulates human-like behavior to avoid bot detection by e-commerce platforms.

    E-commerce sites like Flipkart use sophisticated bot detection systems
    (e.g., PerimeterX, Akamai Bot Manager, DataDome) that analyze:
      - Click timing patterns (bots click at machine speed)
      - Mouse movement trajectories (bots move in straight lines)
      - Typing speed (bots paste entire strings instantly)
      - Scroll behavior (bots jump directly to elements)
      - Session patterns (bots don't pause to "think")
    
    This class implements five strategies to mimic human behavior:
      1. Random delays between actions (1-3.5 seconds)
      2. Character-by-character typing with variable speed
      3. Natural scroll patterns (small increments, not instant jumps)
      4. Random mouse movements to avoid straight-line patterns
      5. Session-level rate limiting (longer pauses every N actions)
    """
    def __init__(self, logger):
        """
        Initialize the human simulator.
        
        Args:
            logger: Logger instance for recording delay decisions
        """
        self.logger = logger
        # Track total actions to implement rate limiting.
        # Every 5 actions, we take a longer pause.
        self.action_count = 0

    def random_delay(self, min_sec=None, max_sec=None):
        """
        Wait a random duration to simulate human thinking/reading time.
        
        This is the most important anti-detection measure. Real humans don't
        click buttons the instant they appear — they read, think, and then act.
        
        Args:
            min_sec: Minimum wait time (defaults to MIN_ACTION_DELAY = 1.0s)
            max_sec: Maximum wait time (defaults to MAX_ACTION_DELAY = 3.5s)
        """
        min_sec = min_sec or MIN_ACTION_DELAY
        max_sec = max_sec or MAX_ACTION_DELAY
        # random.uniform gives a float uniformly distributed between min and max.
        # This is better than random.randint (which only gives integers) because
        # human timing is continuous, not discrete.
        delay = random.uniform(min_sec, max_sec)
        self.logger.debug(f"Human delay: {delay:.2f}s")
        time.sleep(delay)

    def type_like_human(self, element, text):
        """
        Type text character by character with random delays between keystrokes.
        
        Bots typically use element.send_keys("entire string") which pastes
        the text instantly. Real humans type one character at a time with
        variable speed — faster for common sequences, slower for unusual ones.
        
        Args:
            element: Selenium WebElement (input field) to type into
            text: The string to type
        """
        for char in text:
            element.send_keys(char)  # Send one character at a time
            # Variable delay between characters simulates realistic typing speed
            time.sleep(random.uniform(MIN_TYPING_DELAY, MAX_TYPING_DELAY))
        # Small pause after finishing typing (human would review what they typed)
        self.random_delay(0.5, 1.0)

    def scroll_naturally(self, driver, pixels=None):
        """
        Scroll the page in a natural, human-like manner.
        
        Bots use scrollIntoView() which instantly jumps to an element.
        Humans scroll smoothly in multiple small increments.
        
        Args:
            driver: Selenium WebDriver instance
            pixels: Total pixels to scroll (random 200-500 if not specified)
        """
        if pixels is None:
            pixels = random.randint(200, 500)
        # Break the scroll into 3-7 small steps
        steps = random.randint(3, 7)
        step_size = pixels // steps
        for _ in range(steps):
            # Execute JavaScript to scroll by step_size pixels
            driver.execute_script(f"window.scrollBy(0, {step_size});")
            # Tiny delay between scroll steps for smooth animation
            time.sleep(random.uniform(0.05, 0.15))
        # Pause after scrolling (human would read the newly visible content)
        self.random_delay(0.5, 1.0)

    def random_mouse_movement(self, driver):
        """
        Move the mouse cursor to a random position on the page.
        
        Bot detection systems track mouse movement patterns. Real humans move
        their mouse around while reading. This method creates random cursor
        movement to make the session appear more natural.
        
        Note: This is wrapped in try/except because ActionChains may not work
        in all browser configurations (e.g., headless mode). It's a nice-to-have
        enhancement, not a critical feature.
        
        Args:
            driver: Selenium WebDriver instance
        """
        try:
            from selenium.webdriver.common.action_chains import ActionChains
            # Find the body element as a reference point for cursor positioning
            body = driver.find_element("tag name", "body")
            # Get viewport dimensions to stay within visible area
            viewport_width = driver.execute_script("return window.innerWidth;")
            viewport_height = driver.execute_script("return window.innerHeight;")
            # Generate random coordinates (avoiding edges where there are no elements)
            x = random.randint(100, max(101, viewport_width - 100))
            y = random.randint(100, max(101, viewport_height - 100))
            # Move cursor to the random position relative to the body element
            ActionChains(driver).move_to_element_with_offset(body, x, y).perform()
        except Exception:
            pass  # Non-critical; some environments don't support mouse simulation

    def increment_action(self):
        """
        Track action count and apply rate limiting.
        
        Every 5 actions, the agent takes a longer pause (3-6 seconds).
        This mimics a human who periodically pauses to think, read, or
        take a sip of coffee. Without this, even with random delays,
        the consistent action cadence could look suspicious.
        """
        self.action_count += 1
        if self.action_count % 5 == 0:
            # Every 5 actions, take a longer "thinking" break
            self.logger.info(f"Rate limiting pause (action #{self.action_count})")
            self.random_delay(3.0, 6.0)


# -------------
# EXCEL MANAGER
# -------------

def _safe_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None
    

class ExcelManager:
    """
    Handles reading pending tasks from and writing results to the Excel file.
    
    Design decisions:
    1. Uses openpyxl (not pandas) because openpyxl supports cell-level writes
       without rewriting the entire sheet. Pandas would destroy formatting.
    2. Opens and closes the workbook for each operation to prevent file
       corruption if the agent crashes. This is slower but much safer.
    3. Results are written immediately after each task completes, so partial
       progress is preserved even if the agent crashes mid-execution.
    """

    def __init__(self, filepath, logger):
        """
        Args:
        filepath: Absolute path to the Excel file
        logger: Logger instance for recording read/write operations
        """
        self.filepath = filepath
        self.logger = logger

    def read_pending_tasks(self):
        """
        Read all rows with status 'Pending' from the excel file.
        
        Workflow:
        1. Open the workbook in read mode
        2. Iterate through all rows starting from row 2 (row 1 = headers)
        3. Filter rows where the Status column (J) is "Pending"
        4. Convert each matching row into a ReturnTask dataclass instance
        5. Close the workbook and return the list of tasks
        
        Returns:
            list[ReturnTask]: List of pending tasks to process
        """
        # Opent the workbook.
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active      # Get's the first active sheet (we only have one sheet)
        tasks = []

        # Iterate through rows starting from row 2 (skip header row 1).
        # enumerate(..., start=2) makes row_idx match the actual Excel row number,
        # which we need later for writing results back to the correct row.
        # values_only=False returns Cell objects (not just values) — we use this
        # to access cell.value for each column.
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Extract the value from each cell in the row
            values = [cell.value for cell in row]
            # Skip rows with fewer than 11 columns (malformed/empty rows)
            if len(values) < 11:
                continue

            # Check if this row's status is "Pending" (case-insensitive).
            # str(values[9] or '') handles None values safely.
            status = str(values[9] or '').strip()
            if status.lower() != 'pending':
                continue    # Skip non-pending tasks

            # Create a ReturnTask instance from the row data.
            # Each values[N] corresponds to a column (0-indexed):
            #   0=Address, 1=Contact, 2=Product Link, 3=Amount, etc.
            # str(values[N] or '') safely converts None to empty string.
            # float(values[N] or 0) safely converts None to 0.0.
            task = ReturnTask(
                row_number=row_idx,                                       # Excel row for write-back
                address=str(values[0] or ''),                             # Column A
                contact=str(values[1] or ''),                             # Column B
                product_link=str(values[2] or ''),                        # Column C
                amount=float(values[3] or 0),                             # Column D (₹)
                num_products=int(values[4] or 0),                         # Column E
                order_date=str(values[5] or ''),                          # Column F
                order_id=str(values[6] or ''),                            # Column G
                delivery_date=str(values[7] or ''),                       # Column H
                return_window=str(values[8] or ''),                       # Column I
                status=status,                                            # Column J
                platform=str(values[10] or ''),                           # Column K
                refund_id=str(values[11] or '') if len(values) > 11 else None,      # Column L
                return_status=str(values[12] or '') if len(values) > 12 else None,  # Column M
                refund_amount=_safe_float(values[13]) if len(values) > 13 else None, # Column N
                timestamp=str(values[14] or '') if len(values) > 14 else None,      # Column O
                log=str(values[15] or '') if len(values) > 15 else None,            # Column P
            )
            tasks.append(task)

        wb.close()  # Always close to release the file handle
        self.logger.info(f"Found {len(tasks)} pending tasks in Excel")
        return tasks

    def write_result(self, task: ReturnTask, result: ReturnResult):
        """
        Write the return result back to Excel for a specific line item.
        
        This method implements the "per-line-item write-back" requirement:
        each SKU/product gets its own result, not just the order as a whole.
        
        The method opens the workbook, writes to the specific row, saves,
        and closes. This per-operation I/O pattern is intentional:
          - If the agent crashes after processing 3 of 7 tasks, the first 3
            results are safely persisted in the file.
          - A batch approach (write all at end) would lose ALL results on crash.
        
        Args:
            task: The original task (contains row_number for write-back)
            result: The processing result to write
        """
        # Open the workbook fresh for each write (not cached — safer)
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        row = task.row_number  # The exact excel row to update

        # Write result columns (L through P):
        ws.cell(row=row, column=12, value=result.return_id or '')         # Column L: Refund ID
        ws.cell(row=row, column=13, value=result.return_status)           # Column M: Return Status
        ws.cell(row=row, column=14, value=result.refund_amount)           # Column N: Refund Amount
        ws.cell(row=row, column=15,                                       # Column O: Timestamp
                value=datetime.now().isoformat())  # ISO 8601 format for consistency
        ws.cell(row=row, column=16, value=result.log_message)             # Column P: Log/Error

        # Update the Status column (J) based on the outcome:
        if result.success:
            ws.cell(row=row, column=10, value='Done')        # Task completed successfully
        elif 'human review' in result.return_status.lower():
            ws.cell(row=row, column=10, value='Needs Review') # Needs manual intervention

        # Save and close. This persists the changes to disk immediately.
        wb.save(self.filepath)
        wb.close()
        self.logger.info(f"Row {row}: Written result — Status: {result.return_status}, "
                         f"Refund ID: {result.return_id}")


# -----------------------
# FLIPKART RETURN HANDLER
# -----------------------

class FlipkartHandler:
    """
    Handles the complete Flipkart return flow using Selenium WebDriver.
    
    This class encapsulates all Flipkart-specific logic:
      - Login via phone number + OTP
      - Order navigation and lookup
      - Return initiation (reason selection, refund method, confirmation)
      - Return ID and refund amount extraction
      - Screenshot capture for debugging
    
    The class is designed to be replaceable — if Amazon support is needed,
    an AmazonHandler class would implement the same interface (navigate_to_order,
    initiate_return) but with Amazon-specific UI interactions.
    
    Flipkart Return Flow (typical):
      1. Go to My Orders page
      2. Find the specific order by ID
      3. Click "Return" button on the item
      4. Select a return reason (radio button)
      5. Optionally add a comment
      6. Click Continue/Next
      7. Select refund method (original payment / bank account)
      8. Select pickup option (scheduled pickup / self ship)
      9. Click Confirm
      10. Capture the Return ID from the confirmation page
    """

    def __init__(self, driver, human_sim, logger):
        """
        Args:
            driver: Selenium WebDriver instance (Chrome)
            human_sim: HumanSimulator instance for anti-detection delays
            logger: Logger for recording actions and errors
        """
        self.driver = driver
        self.human = human_sim
        self.logger = logger

    def login(self, phone_number, max_retries=MAX_RETRIES):
        """
        Login to Flipkart using phone number + OTP authentication.
        
        Flipkart's login flow:
          1. Navigate to login page
          2. Enter phone number in the input field
          3. Click "Request OTP" button
          4. Wait for OTP to be entered (manually, by calling 9205359199)
          5. Verify login success by checking if URL changed away from /login
        
        OTP handling: In this implementation, OTP must be entered manually.
        A production version would integrate with an SMS gateway API to
        automatically retrieve and enter the OTP.
        
        Args:
            phone_number: The phone number to log in with
            max_retries: Number of login attempts before giving up
            
        Returns:
            bool: True if login succeeded, False if all attempts failed
        """
        for attempt in range(max_retries):
            try:
                self.logger.info(f"Flipkart login attempt {attempt + 1}/{max_retries}")
                # Navigate to the login page
                self.driver.get(FLIPKART_LOGIN_URL)
                self.human.random_delay(2, 4)  # Wait for page to fully render

                # Flipkart sometimes shows a login popup that blocks interaction.
                # Try to close it if it appears.
                try:
                    from selenium.webdriver.common.by import By
                    from selenium.webdriver.support.ui import WebDriverWait
                    from selenium.webdriver.support import expected_conditions as EC

                    # Look for the popup dismiss button (class names may change)
                    close_btn = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                    "button._2KpZ6l._2doB4z"))
                    )
                    close_btn.click()
                    self.human.random_delay()
                except Exception:
                    pass  # No popup appeared — continue normally

                # Navigate to login page again (after closing popup)
                self.driver.get(FLIPKART_LOGIN_URL)
                self.human.random_delay(2, 4)

                # Import Selenium utilities for element location and waiting.
                # These are imported inside the method because they're only needed
                # when actually running the browser (not in dry-run mode).
                from selenium.webdriver.common.by import By
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC

                # The input field doesn't have a placeholder attribute directly.
                # Instead, it's followed by a label/span containing the text "Enter Email/Mobile number".
                # We use XPath to find an input that has a sibling label or span containing "Email" or "Mobile",
                # or fallback to the first text input inside a form.
                phone_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                        "//form[.//button[contains(., 'OTP') or contains(., 'Continue') or contains(., 'CONTINUE') or contains(., 'LOGIN')]]//input[@type='text']"
                    ))
                )
                phone_input.clear()  # Clear any pre-filled value
                # Type the phone number character by character (anti-bot measure)
                self.human.type_like_human(phone_input, phone_number)
                self.human.random_delay()

                # Find and click the "Request OTP" / "Continue" / "LOGIN" button.
                # Using XPath with contains() makes this robust against text variations.
                # The | (pipe) in XPath means OR — matches any of the listed text values.
                continue_btn = self.driver.find_element(
                    By.XPATH,
                    "//button[contains(text(), 'Request OTP') or "
                    "contains(text(), 'CONTINUE') or "
                    "contains(text(), 'Continue') or "
                    "contains(text(), 'LOGIN')]"
                )
                continue_btn.click()
                self.human.random_delay(2, 3)

                # Wait for the OTP input field to appear
                self.logger.info("⏳ Waiting for OTP entry... "
                                 "Please enter OTP manually or call 9205359199 to get it.")

                # Wait up to 120 seconds for the OTP input field to appear.
                # The CSS selector matches various OTP input styles:
                #   - maxlength attribute (OTP fields often limit to 4-6 digits)
                #   - type='number' (numeric-only input)
                #   - class containing 'otp'
                otp_input = WebDriverWait(self.driver, 120).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR,
                        "input[type='text'][maxlength], input[type='number'], "
                        "input[class*='otp']"))
                )

                self.logger.info("🔑 OTP input field found. Enter OTP manually...")

                # Wait up to 180 seconds (3 minutes) for the user to enter OTP
                # and the page to navigate away from the login URL.
                # In production, this would be automated via SMS API integration.
                WebDriverWait(self.driver, 180).until(
                    lambda d: "login" not in d.current_url.lower()
                )
                self.human.random_delay(3, 5)

                # Verify login was successful by checking if we're still on the login page.
                # If the URL no longer contains "login", authentication succeeded.
                if "login" not in self.driver.current_url.lower():
                    self.logger.info("✅ Flipkart login successful!")
                    return True

            except Exception as e:
                self.logger.error(f"Login attempt {attempt + 1} failed: {str(e)}")
                # Wait longer between failed attempts (exponential-ish backoff)
                self.human.random_delay(5, 10)

        self.logger.error("All login attempts failed")
        return False

    def navigate_to_order(self, order_id):
        """
        Navigate to a specific order's details page on Flipkart.
        
        Strategy (tries multiple approaches):
          1. Direct URL with order ID parameter
          2. Search for the order on the orders page
          3. Scroll through the orders page to find it
        
        Args:
            order_id: Flipkart order ID (e.g., "OD337915012166989100")
            
        Returns:
            bool: True if the order was found on the page, False otherwise
        """
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC

            self.logger.info(f"Navigating to order: {order_id}")

            # Approach 1: Go to My Orders page first
            self.driver.get(FLIPKART_ORDERS_URL)
            self.human.random_delay(2, 4)

            # Approach 2: Try direct URL with order ID as a query parameter.
            # This sometimes works on Flipkart to jump directly to the order.
            order_url = f"https://www.flipkart.com/account/orders?orderItemId={order_id}"
            self.driver.get(order_url)
            self.human.random_delay(2, 4)

            # Approach 3: Try the search functionality on the orders page.
            # Not all Flipkart versions have this, so it's wrapped in try/except.
            try:
                # Use a specific placeholder text to avoid matching the global header search bar
                search_input = self.driver.find_element(
                    By.CSS_SELECTOR,
                    "input[placeholder*='orders']"
                )
                search_input.clear()
                self.human.type_like_human(search_input, order_id)
                self.human.random_delay(2, 3)

                # Use XPath to find the specific button for searching orders
                search_btn = self.driver.find_element(
                    By.XPATH,
                    "//button[contains(., 'Search Orders')]"
                )
                search_btn.click()
                self.human.random_delay(2, 4)
            except Exception:
                self.logger.debug("Search not available, scrolling to find order")

            # Final check: look for the order ID text anywhere on the page.
            # XPath //*[contains(text(), 'ORDER_ID')] finds any element containing
            # the order ID string.
            try:
                order_element = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH,
                        f"//*[contains(text(), '{order_id}')]"))
                )
                self.human.scroll_naturally(self.driver)
                self.logger.info(f"✅ Found order {order_id} on page")
                return True
            except Exception:
                self.logger.warning(f"Order {order_id} not found on page")
                return False

        except Exception as e:
            self.logger.error(f"Error navigating to order {order_id}: {str(e)}")
            return False

    def initiate_return(self, task: ReturnTask) -> ReturnResult:
        """
        Initiate the return flow for a specific line item on Flipkart.
        
        This is the core automation method. It follows the Flipkart return flow:
          1. Pre-check: skip if already cancelled/not delivered
          2. Navigate to the order
          3. Click the "Return" button
          4. Select a return reason
          5. Add an optional comment
          6. Click Continue/Next
          7. Select refund method
          8. Select pickup option
          9. Confirm the return
          10. Capture the Return ID and refund amount from the confirmation
        
        PARTIAL-SUCCESS HANDLING:
        Each step is wrapped in its own try/except. If selecting a reason fails,
        the agent still tries to click Continue. If Continue fails, it still
        tries to Confirm. This "best effort" approach maximizes the chance of
        completing the return even if intermediate steps have UI changes.
        
        Args:
            task: The return task containing order details
            
        Returns:
            ReturnResult: The outcome of the return attempt
        """
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC

            self.logger.info(f"Initiating return for Order {task.order_id}")
            self.human.increment_action()  # Track for rate limiting

            # --- PRE-CHECKS: Skip tasks that don't need browser interaction ---
            
            # Check 1: Already cancelled/refunded — no action needed
            if task.return_status and 'cancelled' in task.return_status.lower():
                return ReturnResult(
                    success=True,
                    return_id=task.refund_id,
                    return_status="Already Cancelled & Refunded",
                    refund_amount=task.refund_amount,
                    log_message=f"Already cancelled. Refund ID: {task.refund_id}"
                )

            # Check 2: Not yet delivered — can't return what hasn't arrived
            if task.return_status and 'not yet delivered' in task.return_status.lower():
                return ReturnResult(
                    success=False,
                    return_status="Not yet delivered",
                    log_message="Order is not yet delivered. Cannot initiate return."
                )

            # --- NAVIGATE TO ORDER ---
            if not self.navigate_to_order(task.order_id):
                return ReturnResult(
                    success=False,
                    return_status="Order Not Found",
                    log_message=f"Could not find order {task.order_id} on Flipkart"
                )

            # --- STEP 1: Find and click "Return" button ---
            # Multiple element types checked: <a>, <button>, <span>
            # because Flipkart renders the return action differently depending
            # on the order type and UI version.
            try:
                return_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH,
                        "//a[contains(text(), 'Return')] | "
                        "//button[contains(text(), 'Return')] | "
                        "//span[contains(text(), 'Return')]"
                    ))
                )
                self.human.random_delay()
                return_btn.click()
                self.human.random_delay(2, 4)
            except Exception:
                # No return button = likely out of return window or needs support
                screenshot_path = self._take_screenshot(f"no_return_btn_{task.order_id}")
                return ReturnResult(
                    success=False,
                    return_status="Support Needed",
                    log_message="No return button found. May be out of return window "
                                "or requires manual support.",
                    screenshot_path=screenshot_path
                )

            # --- STEP 2: Select return reason ---
            # Flipkart presents a list of reasons as radio buttons or labels.
            # We try multiple common reason texts in priority order.
            try:
                reasons = [
                    "Product quality issue",
                    "Item not as described",
                    "Wrong item received",
                    "Product damaged",
                    "Quality not as expected"
                ]

                # Try each reason text until one is found on the page
                for reason_text in reasons:
                    try:
                        reason_elem = self.driver.find_element(
                            By.XPATH,
                            f"//span[contains(text(), '{reason_text}')] | "
                            f"//label[contains(text(), '{reason_text}')]"
                        )
                        reason_elem.click()
                        self.human.random_delay()
                        break  # Found and clicked — stop trying other reasons
                    except Exception:
                        continue  # This reason text not found — try next one

                # Fallback: if none of the specific reasons matched, try clicking
                # the first radio button or reason div on the page
                try:
                    radio_btns = self.driver.find_elements(
                        By.CSS_SELECTOR,
                        "input[type='radio'], div[class*='reason']"
                    )
                    if radio_btns:
                        radio_btns[0].click()
                        self.human.random_delay()
                except Exception:
                    pass  # Reason selection is best-effort

            except Exception as e:
                self.logger.warning(f"Could not select return reason: {e}")

            # --- STEP 3: Add optional comment ---
            # Some return flows have a text area for additional comments.
            # This is optional — the return can proceed without it.
            try:
                comment_box = self.driver.find_element(
                    By.CSS_SELECTOR,
                    "textarea, input[type='text'][placeholder*='comment']"
                )
                self.human.type_like_human(comment_box, "Product quality not as expected")
                self.human.random_delay()
            except Exception:
                pass  # No comment box — that's fine

            # --- STEP 4: Click Continue/Next/Submit ---
            try:
                next_btn = self.driver.find_element(
                    By.XPATH,
                    "//button[contains(text(), 'Continue')] | "
                    "//button[contains(text(), 'CONTINUE')] | "
                    "//button[contains(text(), 'Next')] | "
                    "//button[contains(text(), 'Submit')]"
                )
                next_btn.click()
                self.human.random_delay(2, 4)
            except Exception:
                pass  # Button might not exist in this flow variant

            # --- STEP 5: Select refund method ---
            # Options typically: "Original Payment Method" or "Bank Account"
            try:
                refund_options = self.driver.find_elements(
                    By.XPATH,
                    "//span[contains(text(), 'Original Payment')] | "
                    "//span[contains(text(), 'Bank Account')] | "
                    "//label[contains(text(), 'Refund')]"
                )
                if refund_options:
                    refund_options[0].click()  # Select the first available option
                    self.human.random_delay()
            except Exception:
                pass

            # --- STEP 6: Select pickup option ---
            # Options typically: "Schedule Pickup" or "Self Ship"
            try:
                pickup_options = self.driver.find_elements(
                    By.XPATH,
                    "//span[contains(text(), 'Schedule Pickup')] | "
                    "//span[contains(text(), 'Self Ship')] | "
                    "//label[contains(text(), 'pickup')]"
                )
                if pickup_options:
                    pickup_options[0].click()  # Select the first available option
                    self.human.random_delay()
            except Exception:
                pass

            # --- STEP 7: Confirm the return ---
            try:
                confirm_btn = self.driver.find_element(
                    By.XPATH,
                    "//button[contains(text(), 'Confirm')] | "
                    "//button[contains(text(), 'CONFIRM')] | "
                    "//button[contains(text(), 'Submit Return')] | "
                    "//button[contains(text(), 'SUBMIT')]"
                )
                confirm_btn.click()
                self.human.random_delay(3, 5)  # Longer wait for confirmation processing
            except Exception:
                pass

            # --- STEP 8: Capture return ID and refund amount ---
            # After confirmation, the page should show return details.
            # We extract the Return ID and refund amount using regex patterns.
            return_id = None
            refund_amount = None

            try:
                # Get all visible text on the page
                page_text = self.driver.find_element(By.TAG_NAME, "body").text

                # Extract Return/Refund ID using regex.
                # Matches patterns like "Return ID: ABC123" or "Refund #XYZ789"
                import re
                return_id_match = re.search(
                    r'(?:Return|Refund)\s*(?:ID|#|Number)[:\s]*([A-Z0-9]+)',
                    page_text, re.IGNORECASE
                )
                if return_id_match:
                    return_id = return_id_match.group(1)

                # Extract refund amount.
                # Matches patterns like "Refund: ₹1,234.56" or "Amount: 5678"
                amount_match = re.search(
                    r'(?:Refund|Amount)[:\s]*₹?\s*([\d,]+(?:\.\d{2})?)',
                    page_text, re.IGNORECASE
                )
                if amount_match:
                    refund_amount = float(amount_match.group(1).replace(',', ''))

                # Check for success indicators in the page text.
                # If any of these keywords appear, the return was likely successful.
                if any(keyword in page_text.lower() for keyword in
                       ['return initiated', 'return placed', 'successfully',
                        'return request', 'pickup scheduled']):
                    screenshot_path = self._take_screenshot(
                        f"return_success_{task.order_id}")
                    return ReturnResult(
                        success=True,
                        return_id=return_id,
                        return_status="Return Placed",
                        refund_amount=refund_amount or task.amount,
                        log_message=f"Return successfully initiated for order {task.order_id}",
                        screenshot_path=screenshot_path
                    )

            except Exception as e:
                self.logger.warning(f"Could not capture return details: {e}")

            # If we reached here without confirming success, flag for human review.
            # This is the "fail open" approach — we don't mark it as done unless
            # we're SURE it succeeded. Uncertain results get flagged for review.
            screenshot_path = self._take_screenshot(f"review_needed_{task.order_id}")
            return ReturnResult(
                success=False,
                return_id=return_id,
                return_status="Needs human review",
                refund_amount=refund_amount,
                log_message=f"Return flow completed but could not verify success. "
                            f"Screenshot saved for review.",
                screenshot_path=screenshot_path
            )

        except Exception as e:
            # Catch-all for unexpected errors. Take a screenshot for debugging
            # and return a failure result with the error message.
            self.logger.error(f"Error processing return for {task.order_id}: {str(e)}")
            screenshot_path = self._take_screenshot(f"error_{task.order_id}")
            return ReturnResult(
                success=False,
                return_status="Failed",
                log_message=f"Error: {str(e)}",
                screenshot_path=screenshot_path
            )

    def _take_screenshot(self, name):
        """
        Capture a browser screenshot for debugging and review.
        
        Screenshots are saved with a descriptive name and timestamp:
          screenshots/no_return_btn_OD123_20260730_235357.png
        
        This provides visual evidence of what the page looked like when
        the agent encountered an issue — invaluable for debugging.
        
        Args:
            name: Descriptive prefix for the screenshot filename
            
        Returns:
            str or None: Path to the saved screenshot, or None if saving failed
        """
        try:
            os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
            path = os.path.join(SCREENSHOTS_DIR,
                                f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
            self.driver.save_screenshot(path)
            self.logger.info(f"📸 Screenshot saved: {path}")
            return path
        except Exception as e:
            self.logger.warning(f"Could not save screenshot: {e}")
            return None


# ------------------
# BROWSER CONTROLLER
# ------------------
class BrowserController:
    """
    Manages the Chrome browser instance with stealth/anti-detection features.
    
    This class handles browser lifecycle (start/stop) and applies stealth
    configurations to avoid bot detection. It tries two approaches:
    
    1. PREFERRED: undetected-chromedriver (pip install undetected-chromedriver)
       - A modified ChromeDriver that patches common bot-detection fingerprints
       - Automatically handles ChromeDriver version matching
       - Bypasses most PerimeterX/Akamai/DataDome checks
    
    2. FALLBACK: Standard Selenium with manual stealth patches
       - Applied when undetected-chromedriver is not installed
       - Overrides navigator.webdriver JavaScript property
       - Hides automation indicators (extension flags, empty plugins)
       - Less reliable than undetected-chromedriver but still effective
    """

    def __init__(self, logger, headless=False):
        """
        Args:
            logger: Logger instance for recording browser events
            headless: If True, run browser without visible GUI window
                      (useful for server environments, but less stealthy)
        """
        self.logger = logger
        self.headless = headless
        self.driver = None

    def start(self):
        """
        Initialize the browser with stealth settings.
        
        Returns:
            WebDriver: The configured Selenium WebDriver instance
        """
        try:
            # --- Attempt 1: Use undetected-chromedriver ---
            # This is a drop-in replacement for selenium.webdriver.Chrome
            # that automatically patches Chrome to avoid bot detection.
            import undetected_chromedriver as uc
            options = uc.ChromeOptions()
            if self.headless:
                # '--headless=new' uses Chrome's new headless mode (v2),
                # which is harder for sites to detect than the old headless mode.
                options.add_argument('--headless=new')
            # Set a realistic viewport size (1920x1080 = standard Full HD).
            # Unusual sizes (like 800x600) can trigger bot detection.
            options.add_argument('--window-size=1920,1080')
            # Disable the "AutomationControlled" Blink feature flag.
            # This flag is the primary way Chrome exposes automation to websites.
            options.add_argument('--disable-blink-features=AutomationControlled')
            # Set language to English (US) for consistent page content
            options.add_argument('--lang=en-US,en')

            self.driver = uc.Chrome(options=options)
            self.logger.info("✅ Browser started (undetected-chromedriver — stealth mode)")

        except ImportError:
            # --- Attempt 2: Fall back to standard Selenium ---
            # If undetected-chromedriver is not installed, use regular Selenium
            # with manual stealth patches applied via Chrome DevTools Protocol (CDP).
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service

            options = Options()
            if self.headless:
                options.add_argument('--headless=new')
            options.add_argument('--window-size=1920,1080')
            options.add_argument('--disable-blink-features=AutomationControlled')
            # '--no-sandbox': Required in some Linux environments (Docker, CI).
            # Disables Chrome's sandbox security (safe for non-production use).
            options.add_argument('--no-sandbox')
            # '--disable-dev-shm-usage': Prevents Chrome from running out of
            # shared memory in containers with limited /dev/shm.
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--lang=en-US,en')

            # Remove the "Chrome is being controlled by automated test software" infobar.
            # excludeSwitches: removes command-line switches that expose automation.
            # useAutomationExtension: disables the automation extension entirely.
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)

            self.driver = webdriver.Chrome(options=options)

            # --- Apply JavaScript stealth patches via Chrome DevTools Protocol ---
            # These scripts run BEFORE any page's JavaScript, overriding browser
            # properties that bot detectors check.
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    // Override navigator.webdriver: The #1 bot detection signal.
                    // In automated Chrome, navigator.webdriver is true.
                    // In normal Chrome, it's undefined.
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    });
                    
                    // Override navigator.plugins: Normal browsers have plugins
                    // (PDF viewer, etc.). Automated browsers often have 0 plugins.
                    // We fake a list of 5 plugins to appear legitimate.
                    Object.defineProperty(navigator, 'plugins', {
                        get: () => [1, 2, 3, 4, 5]
                    });
                    
                    // Override navigator.languages: Set realistic language prefs.
                    // Some bots forget to set this, leaving it empty.
                    Object.defineProperty(navigator, 'languages', {
                        get: () => ['en-US', 'en']
                    });
                    
                    // Add window.chrome.runtime: Present in real Chrome but
                    // missing in some automated configurations.
                    window.chrome = { runtime: {} };
                """
            })

            self.logger.info("Browser started (standard Selenium with stealth patches)")

        # Set implicit wait: if an element isn't found immediately, Selenium
        # will retry for up to PAGE_LOAD_TIMEOUT seconds before raising an error.
        self.driver.implicitly_wait(PAGE_LOAD_TIMEOUT)
        return self.driver

    def stop(self):
        """
        Close the browser cleanly, releasing all resources.
        
        This is called in a `finally` block to ensure the browser closes
        even if the agent crashes. Unclosed Chrome processes consume memory
        and can cause issues on subsequent runs.
        """
        if self.driver:
            try:
                self.driver.quit()  # quit() closes the browser AND kills the driver process
                self.logger.info("Browser closed")
            except Exception as e:
                self.logger.warning(f"Error closing browser: {e}")


# –––––––––––––––––––––––
# MAIN AGENT ORCHESTRATOR
# –––––––––––––––––––––––

class ReturnAgent:
    """
    Main orchestrator that coordinates the entire return automation workflow.
    
    This is the entry point for the agent. It:
        1. Reads pending tasks from excel
        2. Groups tasks by platform (e.g., Flipkart, Amazon)
        3. For each platform group:
            a. Starts a browser with stealth settings
            b. Logs into the platform
            c. Processes each line item independently (partial-success)
            d. Writes results back to excel after each item
        4. Prints a summary of all results
        
    The "independence" of line items is a critical design choice:
    if Item A fails (e.g., out of return window), the agent continues
    processing Items B and C. It never abandons an entire order because
    one item couldn't be returned.
    """
    
    def __init__(self, dry_run=False, headless=False, target_order_id=None):
        """
        Args:
            dry_run: If True, simulate everything without opening a browser.
                     Useful for testing the excel reading/writing pipeline.
            headless: If True, run the browser without a visible window.
            target_order_id: If set, only process this specific order ID.
        """
        self.logger = setup_logging()       # Initialize logging first
        self.dry_run = dry_run
        self.headless = headless
        self.target_order_id = target_order_id
        self.excel_manager = ExcelManager(EXCEL_FILE, self.logger)
        # Accumulates (task, result) tuples for the final summary report
        self.results_summary = []

    def run(self):
        """
        Execute the main return automation workflow.
        
        This is the top-level orchestration method that coordinates all
        components. The workflow is:
        
        Step 1: Read pending tasks from excel
        Step 2: Group tasks by platform
        Step 3: Process each platform group (browser automation)
        Step 4: Print Summary of all results
        """
        # Print startup banner with config
        self.logger.info("=" * 60)
        self.logger.info("RETURN AUTOMATION AGENT - STARTED")
        self.logger.info(f"Dry run: {self.dry_run}")
        self.logger.info(f"Target order: {self.target_order_id or 'ALL'}")
        self.logger.info("=" * 60)

        # Step 1: Read all pending tasks from excel spreadsheet
        tasks = self.excel_manager.read_pending_tasks()     # returns a ReturnTask instance
        if not tasks:
            self.logger.info("No pending tasks found. Exiting.")

        # If a specific order ID was requested, filter to just that order
        if self.target_order_id:
            tasks = [t for t in tasks if t.order_id == self.target_order_id]
            self.logger.info(f"Filtered to {len(tasks)} tasks for order {self.target_order_id}")

        # Step 2: Group tasks by platform
        platform_groups = {}
        for task in tasks:
            platform = task.platform.strip()
            if platform not in platform_groups:
                platform_groups[platform] = []
            platform_groups[platform].append(task)

        self.logger.info(f"Task groups: {', '.join(f'{k}: {len(v)}' for k, v in platform_groups.items())}")

        # Step 3: Process each platform group
        for platform, platform_tasks in platform_groups.items():
            self.logger.info(f"\n{'=' * 40}")
            self.logger.info(f"Processing platform: {platform}")
            self.logger.info(f"Tasks: {len(platform_tasks)}")
            self.logger.info(f"{'=' * 40}")

            if platform.lower() == 'flipkart':
                self._process_flipkart_tasks(platform_tasks)
            else:
                # Unsupported platform - flag all tasks for manual handling.
                # This is extensible: add elif blocks for Amazon, Meesho, etc.
                self.logger.warning(f"Platform '{platform}' not yet supported. "
                                    f"Flagging {len(platform_tasks)} tasks for review.")
                for task in platform_tasks:
                    result = ReturnResult(
                        success=False,
                        return_status="Platform not supported",
                        log_message=f"Platform '{platform}' automation not implemented"
                    )
                    self._record_result(task, result)

        # Step 4: Print final summary
        self._print_summary()

    def _process_flipkart_tasks(self, tasks):
        """
        Process all Flipkart return tasks in a single browser session.
        
        In dry-run mode, simulates processing without opening a browser.
        In live mode: starts browser -> logs in -> processes each task -> closes.
        
        The try/finally pattern ensures the browser is ALWAYS closed,
        even if the agent crashes during task processing.
        
        Args:
            tasks: list of ReturnTask instances for Flipkart orders
        """
        if self.dry_run:
            self.logger.info("Dry run mode - simulating Flipkart returns")
            for task in tasks:
                self._dry_run_task(task)
            return

        # Initialize browser and helper objects
        browser = BrowserController(self.logger, headless=self.headless)
        driver = browser.start()
        human = HumanSimulator(self.logger)
        handler = FlipkartHandler(driver, human, self.logger)

        try:
            # Authenticate with Flipkart
            login_success = handler.login(FLIPKART_PHONE)
            if not login_success:
                # If login fails after all retries, flag ALL tasks and abort.
                self.logger.error("Could not login to Flipkart. Flagging all tasks.")
                for task in tasks:
                    result = ReturnResult(
                        success=False,
                        return_status="Login failed",
                        log_message="Could not login to Flipkart"
                    )
                    self._record_result(task, result)
                return

            # Process each task INDEPENDENTLY - this is the partial-success handling.
            for i, task in enumerate(tasks):
                self.logger.info(f"\n--- Task {i + 1}/{len(tasks)} ---")
                self.logger.info(f"Order: {task.order_id} | Amount: ₹{task.amount}")

                # Pre-check: some tasks already have a status from previous runs.
                if task.return_status and task.return_status.lower() not in ['', 'none', 'pending']:
                    if 'cancelled' in task.return_status.lower():
                        result = ReturnResult(
                            success=True,
                            return_id=task.refund_id,
                            return_status="Already Cancelled & Refunded",
                            refund_amount=task.refund_amount,
                            log_message=f"Skipped: {task.return_status}"
                        )
                    elif 'not yet delivered' in task.return_status.lower():
                        result = ReturnResult(
                            success=False,
                            return_status="Not yet delivered",
                            log_message="Order not yet delivered"
                        )
                    elif 'support' in task.return_status.lower():
                        result = ReturnResult(
                            success=False,
                            return_status="Support Needed",
                            log_message=task.log or "Manual support needed"
                        )
                    else:
                        # Unknown status - try the return flow
                        result = handler.initiate_return(task)
                else:
                    # No pre-existing status - initiate the full return flow
                    result = handler.initiate_return(task)

                # Write result to excel immediately (crash-safe)
                self._record_result(task, result)
                # Pause between tasks to avoid rate limiting
                human.random_delay(2, 5)

        except Exception as e:
            self.logger.error(f"Critical error: {str(e)}")
        finally:
            # ALWAYS close the browser, even if an exception occured.
            browser.stop()

    def _dry_run_task(self, task: ReturnTask):
        """
        Simulate processing a task without any browser interaction.
        
        This is used in --dry-run mode to test:
          - Excel reading logic
          - Task routing logic (cancelled → skip, not delivered → skip, etc.)
          - Result classification
          - Summary generation
        
        No files are modified, no browser is opened.
        
        Args:
            task: The task to simulate
        """
        self.logger.info(f"[DRY RUN] Order: {task.order_id} | "
                         f"Amount: ₹{task.amount} | "
                         f"Platform: {task.platform} | "
                         f"Current Status: {task.return_status}")

        # Route the task based on its pre-existing status
        if task.return_status:
            status = task.return_status.lower()
            if 'cancelled' in status:
                result = ReturnResult(
                    success=True,
                    return_id=task.refund_id,
                    return_status="Already Cancelled & Refunded",
                    refund_amount=task.refund_amount,
                    log_message=f"[DRY RUN] Skipped — already cancelled"
                )
            elif 'not yet delivered' in status:
                result = ReturnResult(
                    success=False,
                    return_status="Not yet delivered",
                    log_message="[DRY RUN] Cannot return — not yet delivered"
                )
            elif 'support' in status:
                result = ReturnResult(
                    success=False,
                    return_status="Support Needed",
                    log_message=f"[DRY RUN] Needs manual support: {task.log}"
                )
            else:
                result = ReturnResult(
                    success=False,
                    return_status="Simulated — Would initiate return",
                    log_message=f"[DRY RUN] Would navigate to Flipkart and initiate return"
                )
        else:
            result = ReturnResult(
                success=False,
                return_status="Simulated — Would initiate return",
                log_message=f"[DRY RUN] Would navigate to Flipkart and initiate return "
                            f"for order {task.order_id}"
            )

        # Add to summary (but don't write to Excel in dry-run mode)
        self.results_summary.append((task, result))
        self.logger.info(f"  → Result: {result.return_status}")

    def _record_result(self, task: ReturnTask, result: ReturnResult):
        """
        Record a result and write it to excel (unless in dry-run mode).
        
        This is the single point where results are persisted. Having one
        method for this ensures consistent behavior across all code paths.
        
        Args:
            task: The original task
            results: The processing result
        """
        self.results_summary.append((task, result))

        if not self.dry_run:
            try:
                self.excel_manager.write_result(task, result)
            except Exception as e:
                self.logger.error(f"Failed to write result to excel: {e}")
    
    def _print_summary(self):
        """
        Print a formatted summary of all processed tasks.
        
        Shows total/successful/failed counts and individual task outcomes.
        Uses emoji icons (✅,❌) for quick visual scanning."""
        self.logger.info("\n" + "=" * 60)
        self.logger.info("RETURN AUTOMATION - SUMMARY")
        self.logger.info("=" * 60)

        total = len(self.results_summary)
        successful = sum(1 for _, r in self.results_summary if r.success)
        failed = sum(1 for _, r in self.results_summary if not r.success)

        self.logger.info(f"Total tasks processed: {total}")
        self.logger.info(f"Successful: {successful}")
        self.logger.info(f"Failed/Needs Review: {failed}")
        self.logger.info("")

        # Print each task result with status icon and details
        for task, result in self.results_summary:
            status_icon = "✅" if result.success else "❌"
            self.logger.info(f"  {status_icon} Order {task.order_id} | "
                             f"₹{task.amount} | {result.return_status}")
            if result.log_message:
                self.logger.info(f"     └─ {result.log_message}")

        self.logger.info("=" * 60)


# –––––––––––––––
# CLI ENTRY POINT
# –––––––––––––––

def main():
    """
    Parse command-line arguments and launch the return automation agent.
    
    Supported flags:
        --dry-run: Simulate everyting without opening a browser or modifying excel
        --headless: Run the browser without a visible window
        --order-id: Process only a specific order (useful for debugging)

    Examples:
        python3 return_agent.py --dry-run
        python3 return_agent.py
        python3 return_agent.py --order-id OD123456
        python3 return_agent.py --headless
    """
    # argparse creates a cli with help text, type checking, and default values
    parser = argparse.ArgumentParser(
        description='Return Automation Agent - Automates product on Flipkart'
    )
    parser.add_argument('--dry-run', action='store_true',
                        help='Run without browser interaction (simulation mode)')
    parser.add_argument('--headless', action='store_true',
                        help='Run browser in headless mode')
    parser.add_argument('--order-id', type=str, default=None,
                        help='Process a specific order ID only')

    # Parse the arguments from sys.argv
    args = parser.parse_args()

    # Create and run the agent with the parsed config
    agent = ReturnAgent(
        dry_run=args.dry_run,
        headless=args.headless,
        target_order_id=args.order_id
    )
    agent.run()


if __name__ == '__main__':
    main()